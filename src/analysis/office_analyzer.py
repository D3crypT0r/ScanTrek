from pptx import Presentation
from openpyxl import load_workbook
import olefile
import zipfile

class OfficeDissector:
    def analyze(self, file_path):
        if file_path.endswith('.docx'):
            return self.process_docx(file_path)
        elif file_path.endswith('.xlsx'):
            return self.process_xlsx(file_path)
        elif file_path.endswith('.pptx'):
            return self.process_pptx(file_path)
        else:
            return self.process_ole(file_path)

    def process_docx(self, path):
        text = []
        with zipfile.ZipFile(path) as doc:
            xml_content = doc.read('word/document.xml')
            text.extend(self.extract_xml_text(xml_content))
        return {'type': 'docx', 'content': text}

    def process_xlsx(self, path):
        wb = load_workbook(path)
        return {
            'sheets': [{
                'name': sheet.title,
                'data': [[cell.value for cell in row] for row in sheet.iter_rows()]
            } for sheet in wb.worksheets]
        }

    def process_ole(self, path):
        with olefile.OleFileIO(path) as ole:
            return {
                'metadata': ole.get_metadata(),
                'streams': ole.listdir()
            }
